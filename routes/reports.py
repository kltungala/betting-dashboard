from io import BytesIO
from decimal import Decimal

from flask import Blueprint, render_template, request, send_file
from openpyxl import Workbook
from sqlalchemy import func, select

from extensions import db
from models import Bet, Bettor, Fight


reports_bp = Blueprint("reports", __name__, url_prefix="/reports")


@reports_bp.get("/")
def index():
    fights = db.session.scalars(select(Fight).order_by(Fight.fight_date.desc())).all()
    bettors = db.session.scalars(select(Bettor).order_by(Bettor.name)).all()
    selected_fight_ids = {
        int(fight_id) for fight_id in request.args.getlist("fight_id") if fight_id.isdigit()
    }
    if not selected_fight_ids and request.args.get("filters_submitted") != "1":
        selected_fight_ids = {fight.id for fight in fights if fight.status == "settled"}

    player_totals = []
    for bettor in bettors:
        total = sum(
            (
                bet.net_profit
                for bet in bettor.bets
                if bet.fight_id in selected_fight_ids and bet.status in {"won", "lost"}
            ),
            Decimal("0.00"),
        )
        player_totals.append({"bettor": bettor, "total": total})

    commission_total = db.session.scalar(
        select(func.coalesce(func.sum(Bet.commission_amount), 0)).where(Bet.status == "won")
    )
    return render_template(
        "reports/index.html",
        fights=fights,
        bettors=bettors,
        commission_total=commission_total,
        selected_fight_ids=selected_fight_ids,
        player_totals=player_totals,
    )


@reports_bp.get("/export")
def export_excel():
    workbook = Workbook()
    fight_sheet = workbook.active
    fight_sheet.title = "Fight Summary"
    fight_sheet.append(["Date", "Fighter A", "Fighter B", "Status", "Winner", "Side A", "Side B", "Difference"])
    for fight in db.session.scalars(select(Fight).order_by(Fight.fight_date.desc())).all():
        fight_sheet.append([
            fight.fight_date,
            fight.fighter_a,
            fight.fighter_b,
            fight.status,
            fight.winner_side or "",
            float(fight.side_a_total),
            float(fight.side_b_total),
            float(fight.balance_difference),
        ])

    bettor_sheet = workbook.create_sheet("Bettor Summary")
    bettor_sheet.append(["Bettor", "Phone", "Wins", "Losses", "Outstanding Balance"])
    for bettor in db.session.scalars(select(Bettor).order_by(Bettor.name)).all():
        wins = sum(1 for bet in bettor.bets if bet.status == "won")
        losses = sum(1 for bet in bettor.bets if bet.status == "lost")
        bettor_sheet.append([bettor.name, bettor.phone or "", wins, losses, float(bettor.balance)])

    profit_sheet = workbook.create_sheet("Profit Report")
    profit_sheet.append(["Fight", "Bettor", "Gross Winnings", "Commission", "Net Profit"])
    for bet in db.session.scalars(select(Bet).where(Bet.status == "won").order_by(Bet.id)).all():
        profit_sheet.append([
            f"{bet.fight.fighter_a} vs {bet.fight.fighter_b}",
            bet.bettor.name,
            float(bet.gross_winnings),
            float(bet.commission_amount),
            float(bet.net_profit),
        ])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = 20

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="betting-dashboard-report.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
